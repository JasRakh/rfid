import serial
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import pytz

# Настройка последовательного порта
ser = serial.Serial('/dev/tty.usbmodem212101', 9600)
ser.timeout = 2

# Настройка часового пояса Ташкента
tz = pytz.timezone('Asia/Tashkent')

# Генерация имени файла с текущей датой
current_time_title = datetime.now(tz).strftime("%H:%M:%S")
current_date_title = datetime.now(tz).strftime("%Y-%m-%d")
file_name = f"RFID_Records_{current_date_title}-{current_time_title}.xlsx"

# Создание или открытие Excel файла
try:
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RFID Records"
    # Заголовки таблицы
    sheet.append(["UID", "Name", "University ID", "Time In", "Time Out"])

print(f"Saving data to {file_name}")
print("Listening for data...")


def update_or_insert_record(uid, name, university_id):
    # Получаем текущее время
    current_time = datetime.now(tz).strftime("%H:%M:%S")

    # Ищем строку с UID и пустым Time Out
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        if row[0].value == uid and row[4].value is None:  # Найти строку с UID и пустым Time Out
            row[4].value = current_time  # Проставить Time Out
            print(f"Updated Time Out for {uid}")
            return

    # Если UID не найден или Time Out уже заполнен, создаём новую строку с Time In
    sheet.append([uid, name, university_id, current_time, None])
    print(f"Added Time In for {uid}")

try:
    while True:
        data = ser.readline().decode('utf-8').strip()
        if data:
            print(f"Received: {data}")
            # Разбиваем строку на UID, Name, University ID
            parts = data.split(',')
            if len(parts) == 3:
                uid, name, university_id = parts
            else:
                uid, name, university_id = parts[0], "Unknown", "Unknown"

            # Обновляем или добавляем запись
            update_or_insert_record(uid, name, university_id)

            # Сохранение файла
            workbook.save(file_name)
except KeyboardInterrupt:
    print("Stopped by user")
finally:
    ser.close()
    workbook.save(file_name)
    print(f"Serial port closed and data saved to {file_name}.")
